"""
filter_leads.py — Post-processing script to remove big brands/malls
from an already-collected delhi_leads.xlsx file.

Run this AFTER scraper.py finishes:
    python filter_leads.py
    python filter_leads.py --input delhi_leads.xlsx --output delhi_leads_clean.xlsx
"""

import os
import sys
# Add project root to sys.path
root_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if root_dir not in sys.path:
    sys.path.append(root_dir)
import pandas as pd
from config import config

def _is_platform_website(url: str) -> bool:
    """Check if the URL belongs to a known platform."""
    if not isinstance(url, str) or url == "N/A":
        return False
    url_lower = url.lower()
    return any(domain in url_lower for domain in config.PLATFORM_DOMAINS)


def _score_lead(website: str, rating: float) -> str:
    """Assign priority based on website presence and rating."""
    if not isinstance(website, str) or website == "N/A":
        return "🔥 High"
    if _is_platform_website(website):
        if rating and rating >= config.HIGH_RATING_THRESHOLD:
            return "🟡 Medium"
        return "🟢 Low"
    return None # Signals this should be filtered out


def is_local_business(name: str, reviews, website, rating, phone="N/A") -> bool:
    """Return True if the listing is a small/local business and doesn't have a personal website."""
    if not isinstance(name, str):
        return True

    name_lower = name.lower()

    # Check brand blacklist
    for brand in config.BRAND_BLACKLIST:
        if brand.lower() in name_lower:
            return False

    # Check review count cap
    try:
        rev = int(reviews)
        if rev > config.MAX_REVIEWS_FOR_LOCAL:
            return False
    except (ValueError, TypeError):
        pass

    # NEW: Strict Personal Website Filter
    if isinstance(website, str) and website != "N/A":
        if not _is_platform_website(website):
            return False

    # NEW: Contact Details (Phone) Requirement Filter
    if getattr(config, 'REQUIRE_PHONE_NUMBER', False):
        if not isinstance(phone, str) or phone == "N/A":
            return False

    return True


def filter_excel(input_file: str, output_file: str) -> None:
    print(f"📂 Reading: {input_file}")
    df = pd.read_excel(input_file)
    total_before = len(df)
    print(f"   Total rows before filter: {total_before}")

    # Apply filter
    mask = df.apply(
        lambda row: is_local_business(
            row.get("Name", ""), 
            row.get("Total Reviews"), 
            row.get("Website", "N/A"),
            row.get("Rating"),
            phone=row.get("Phone", "N/A")
        ),
        axis=1
    )
    df_clean = df[mask].copy()

    # Re-calculate Priority
    df_clean["Priority"] = df_clean.apply(
        lambda row: _score_lead(row.get("Website", "N/A"), row.get("Rating")),
        axis=1
    )
    
    # ── Priority Sorting ─────────────────────────────────────────────
    # Sort: High (1) -> Medium (2) -> Low (3)
    priority_map = {"🔥 High": 1, "🟡 Medium": 2, "🟢 Low": 3}
    df_clean["_priority_sort"] = df_clean["Priority"].map(priority_map).fillna(99)
    df_clean = df_clean.sort_values(by=["_priority_sort", "Name"]).drop(columns=["_priority_sort"]).reset_index(drop=True)

    removed = total_before - len(df_clean)

    print(f"   Removed (brands/malls/chains): {removed}")
    print(f"✅ Remaining local business leads (sorted by Priority): {len(df_clean)}")

    # Save cleaned file
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_clean.to_excel(writer, index=False, sheet_name="Delhi Leads")
        ws = writer.sheets["Delhi Leads"]

        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for col_cells in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

    print(f"💾 Saved clean file: {output_file}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Filter big brands from leads Excel")
    parser.add_argument("--input",  default="data/delhi_leads.xlsx",       help="Input Excel file")
    parser.add_argument("--output", default="data/delhi_leads_clean.xlsx", help="Output cleaned file")
    args = parser.parse_args()
    filter_excel(args.input, args.output)
